import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
from openpyxl import Workbook
import pandas as pd
import lxml


def convert_xls_to_xlsx(input_file):
    if input_file.endswith('.xls'):

        try:
            html_file = input_file.endswith('.xls')
            tables = pd.read_html(html_file,flavor='lxml')
            df = tables[0]
            xlsx_file = 'output_file.xlsx'
            df.to_excel(xlsx_file, index=False, engine='openpyxl')
            print(f"HTML .xls file converted to {xlsx_file}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to convert {input_file}: {str(e)}")
            return None

    return input_file

def findData(filename, Data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cleaned_value = str(cell.value).strip()
                if Data in cleaned_value:
                    end_index = cleaned_value.index(Data) + len(Data)
                    if end_index == len(Data):
                        next_cell_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                        return next_cell_value

    return "'Data' not found with ending index 10 in the sheet."


def find_course_grades(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    course_data = []

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and 'Course No' in str(cell.value).strip():
                course_row = cell.row
                course_column = cell.column
                point_column = course_column + 4
                sem_value = sheet.cell(row=course_row - 1, column=course_column).value

                year = ""
                for r in range(course_row - 1, 0, -1):
                    year_value = sheet.cell(row=r, column=course_column).value
                    if year_value and (
                            'Diploma First Year' in str(year_value) or 'Diploma Second Year' in str(year_value) or 'Advanced Diploma' in str(year_value)):
                        year = str(year_value).strip()
                        break

                for r in range(course_row + 1, sheet.max_row + 1):
                    course_value = sheet.cell(row=r, column=course_column).value

                    point_value = sheet.cell(row=r, column=point_column).value

                    if point_value is None:
                        point_value = sheet.cell(row=r, column=point_column + 1).value
                    else:
                        point_value = str(point_value).strip()

                    if "Sem" in str(course_value):
                        break

                    if course_value is not None:
                        course_data.append([
                            year,
                            str(sem_value).strip() if sem_value else "",
                            str(course_value).strip(),
                            point_value
                        ])

    return course_data



def create_summary_excel(folder_path, summary_filename):
    summary_wb = Workbook()
    if "Sheet" in summary_wb.sheetnames:
        summary_wb.remove(summary_wb["Sheet"])

    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith('.xlsx') or filename.endswith('.xlsm') or filename.endswith('.xls'):
                input_filename = os.path.join(root, filename)

                if filename.endswith('.xls'):
                    input_filename = convert_xls_to_xlsx(input_filename)
                    if input_filename is None:
                        continue

                if "Summary Sheet.xlsx" in input_filename:
                    continue

                student_no = findData(input_filename, 'Student No')
                if student_no is None:
                    continue

                if student_no in summary_wb.sheetnames:
                    student_sheet = summary_wb[student_no]
                else:
                    student_sheet = summary_wb.create_sheet(title=student_no)

                results = [
                    ['Student No', student_no],
                    ['Student Name', findData(input_filename, 'Student Name')],
                    ['Gender', findData(input_filename, 'Gender')],
                    ['Department', findData(input_filename, 'Department')],
                    ['Specialization', findData(input_filename, 'Specialization')],
                    ['Birth Date', findData(input_filename, 'Birth Date')],
                    ['Probation', findData(input_filename, 'Probation')],
                    [''],
                    ['Year', 'Department', 'Course No', 'Point']
                ]

                for row in results:
                    student_sheet.append(row)

                course_data = find_course_grades(input_filename)
                for year, department, course, point in course_data:
                    student_sheet.append([year, department, course, point])

                print(f"Processed: {input_filename}")

    summary_wb.save(summary_filename)


def exit_app():
    root.destroy()


def select_input_folder():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        input_folder_var.set(folder_selected)
    else:
        print("No folder selected.")


def select_output_folder():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        output_folder_var.set(folder_selected)
    else:
        print("No folder selected.")


def process_files():
    folder_path = input_folder_var.get()
    output_file_name = output_file_name_var.get()
    file_extension = output_extension_var.get()
    output_folder = output_folder_var.get()

    if not folder_path:
        messagebox.showwarning("Input Error", "Please select an input folder.")
        return

    if not output_folder:
        messagebox.showwarning("Input Error", "Please select an output folder.")
        return

    if not output_file_name:
        messagebox.showwarning("Input Error", "Please enter a name for the summary file.")
        return

    if not file_extension.startswith("."):
        file_extension = "." + file_extension

    if file_extension not in [".xlsx", ".xls", ".xlsm"]:
        messagebox.showwarning("Input Error", "Please enter a valid file extension (e.g., .xlsx, .xls, .xlsm).")
        return

    summary_filename = os.path.join(output_folder, f"{output_file_name}{file_extension}")
    create_summary_excel(folder_path, summary_filename)
    messagebox.showinfo("Success", f"Summary sheet created successfully at:\n{summary_filename}")


root = tk.Tk()
root.title("Join Excel Sheet Application")
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = 650
window_height = 400
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)
root.geometry(f'{window_width}x{window_height}+{x}+{y}')

input_folder_var = tk.StringVar()
output_folder_var = tk.StringVar()
output_file_name_var = tk.StringVar()
output_extension_var = tk.StringVar()

tk.Label(root, text="Read Transcript", fg="blue", font=("arial", 20)).place(x=200, y=20)

tk.Label(root, text="Select Input Folder:", font=("arial", 14)).place(x=10, y=70)
tk.Entry(root, textvariable=input_folder_var, font=("arial", 14)).place(x=220, y=72)
tk.Button(root, text="Browse", command=select_input_folder, font=("arial", 14)).place(x=475, y=65)

tk.Label(root, text="Select Output Folder:", font=("arial", 14)).place(x=10, y=130)
tk.Entry(root, textvariable=output_folder_var, font=("arial", 14)).place(x=220, y=132)
tk.Button(root, text="Browse", command=select_output_folder, font=("arial", 14)).place(x=475, y=125)

tk.Label(root, text="Enter Output File Name:", font=("arial", 14)).place(x=10, y=190)
tk.Entry(root, textvariable=output_file_name_var, font=("arial", 14)).place(x=220, y=192)

tk.Label(root, text="Enter File Extension:", font=("arial", 14)).place(x=10, y=250)
tk.Entry(root, textvariable=output_extension_var, font=("arial", 14)).place(x=220, y=252)
tk.Label(root, text="Example: .xlsx / .xls / .xlsm", font=("arial", 10)).place(x=450, y=252)

tk.Button(root, text="Create Excel File", command=process_files, font=("arial", 14, "bold")).place(x=140, y=340)
tk.Button(root, text="Exit", command=exit_app, font=("arial", 14, "bold")).place(x=350, y=340)

root.mainloop()